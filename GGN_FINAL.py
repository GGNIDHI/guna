#!/usr/bin/env python
# coding: utf-8

# In[13]:


def comp1():
    kk=0
    a1=0
    for j in range(13,26):
        for k in range(2,3):
            c5=sheet.cell(row=j,column=k)
            if(c5.value>lst[kk]):
                a1=a1+1
            kk=kk+1
    return a1
    
    
def comp2():
    ll=0
    a2=0
    for j in range(13,26):
        for k in range(3,4):
            c5=sheet.cell(row=j,column=k)
            if(c5.value>lst[ll]):
                a2=a2+1
            ll=ll+1
    return a2
    
def comp3():
    mm=0
    a3=0
    for j in range(13,26):
        for k in range(4,5):
            c5=sheet.cell(row=j,column=k)
            if(c5.value>lst[mm]):
                a3=a3+1
            mm=mm+1
    return a3
        
        
        
def company_evaluate(script,r):
    comp=[]
    t=int(r)+1
    k1=yff.Ticker(script)
    b1=yf.get_balance_sheet(script).transpose()
    e1=yf.get_income_statement(script).transpose()
    i1=e1.iloc[0]['totalRevenue']
    f1=e1.iloc[0]['grossProfit']
    g1=e1.iloc[0]['ebit']
    h1=e1.iloc[0]['netIncome']
    d1=b1.iloc[0]['totalAssets']
    c1=b1.iloc[0]['totalLiab']
    j1=(d1-c1)
    m1=k1.info['marketCap']
    if(m1==None):
        m1=0
    n1=k1.info['freeCashflow']
    if(n1==None):
        n1=0
    comp.extend((i1,f1,g1,h1,d1,c1,j1,m1,n1))
    
    c2=sheet.cell(row=1,column=t)
    c2.value=script
    p=0
    for j in range(2,11):
        i=t
        c1 = sheet.cell(row = j,column = i)
        c1.value=comp[p]
        p=p+1
    
    
    gpm=float(f"{(abs(comp[1]/comp[0])*100):.2f}")
    ebitm=float(f"{(abs(comp[2]/comp[0])*100):.2f}")
    npm=float(f"{(abs(comp[3]/comp[0])*100):.2f}")
    roe=float(f"{(abs(comp[3]/comp[6])*100):.2f}")
    roa=float(f"{(abs(comp[3]/comp[4])*100):.2f}")
    d_e=float(f"{(abs(comp[5]/comp[6])*100):.2f}")
    k=k1.info['currentRatio']
    if(k==None):
        k=0
    cr=float(f"{k:.2f}")
    p_b=float(f"{abs(comp[7]/comp[6]):.2f}")
    p_e=float(f"{abs(comp[7]/comp[3]):.2f}")
    p_s=float(f"{abs(comp[7]/comp[0]):.2f}")
    p_cf=float(f"{abs(comp[7]/comp[8]):.2f}") if int(comp[8])!=0 else 0
    dr1=k1.info['dividendRate']
    if(dr1==None):
        dr1=0
    dr=float(f"{dr1:.2f}")
    dy1=k1.info['dividendYield']
    if(dy1==None):
        dy1=0
    dy=float(f"{dy1:.2f}")
    comp.extend((gpm,ebitm,npm,roe,roa,d_e,cr,p_b,p_e,p_s,p_cf,dr,dy))
    
    for j in range(13,26):
        i=t
        c3=sheet.cell(row=j,column=i)
        c3.value=comp[p]
        p=p+1

def comp_calc():
    for Y in range(1,2):
        cout=0
        W=5
        m=12
        for Q in range(13,26):
            c1 = sheet.cell(row = Q,column = W)
            m=m+1
            for s in range(2,5):
                c2=sheet.cell(row=m,column=s)
                cout=cout+c2.value
            c1.value=float(f"{(cout/3):.2f}")
            lst.append(c1.value)
            cout=0

def comp_calc1():
    for Y in range(1,2):
        cout=0
        W=4
        m=12
        for Q in range(13,26):
            c1 = sheet.cell(row = Q,column = W)
            m=m+1
            for s in range(2,4):
                c2=sheet.cell(row=m,column=s)
                cout=cout+c2.value
            c1.value=float(f"{(cout/3):.2f}")
            lst.append(c1.value)
            cout=0

import yahoo_fin.stock_info as yf
import yfinance as yff
comp=[]
import openpyxl
sectors=["Sales","Gross Profit","EBIT","Net Profit","Assets","Liabilities","Equity","Capitalization","FreeCashFlow"]
sectors1=["Gross Profit Margin","EBIT Margin","Net Profit Margin","ROE","ROA","D/E","Current Ratio","P/B","P/E","P/S","P/CF","Dividend Yield","Dividend Growth"]

wb = openpyxl.Workbook()
sheet = wb.active
k=0
l=0
for i in range(1,2):
    for j in range(2,11):
        c1 = sheet.cell(row = j,column = i)
        c1.value=sectors[k]
        k=k+1
s=0
for i in range(1,2):
    for j in range(13,26):
        c3=sheet.cell(row=j,column=i)
        c3.value=sectors1[s]
        s=s+1

print("Enter how many companies you have placed....")
noc=int(input())
print("Enter the company names as per their share market symbols..seperated by ,")
companies=input()
q=[x for x in companies.split(",")]
import pandas as pd
dfc1 = pd.read_csv('Copy of Yahoo Ticker Symbols - September 2017.csv', header=None, index_col=0, squeeze=True).to_dict()
aabb=dfc1[1]
count=0
lst=[]
r=[]
for i in q:
    if i in aabb:
        r.append(aabb[i])
        count=count+1
        company_evaluate(i,count)
    else:
        print("Please provide the correct company name for ",i)
if(noc==3):
    comp_calc()
    aa=comp1()
    ab=comp2()
    ac=comp3()
    if((aa>ab) and (aa>ac)):
        print(r[0]+"is good comp")
    elif((ab>aa) and (ab>ac)):
        print(r[1]+"is good company")
    elif((ac>aa) and (ac>ab)):
        print(r[3]+"is good comp")
    
elif(noc==2):
    comp_calc1()
    aa=comp1()
    print(aa)
    ab=comp2()
    print(ab)
    
    if(aa>ab):
        print(r[0]+"is good comp")
    elif(ab>aa):
        print(r[1]+"is good comp")
    elif(aa==ab):
        print("Voila you got the golden egg you can chosse any of the two either",r[0],"or",r[1])
#wb.save("fina2.xlsx")



# In[ ]:





# In[12]:


3


# In[ ]:




